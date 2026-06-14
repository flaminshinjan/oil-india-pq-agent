"""Deterministic, pandas-backed reads over the real OIL spreadsheets.

The agents use vector search (Chroma) for narrative content. But the
*morning brief* — the headline insight a CXO sees on load — needs the same
numbers, every time, with full precision. So we also have this module:
straight pandas reads of the canonical Excel files, returning typed values
the agents can quote without hallucination.

Cached on first call; the data is static within a demo run.
"""
from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import openpyxl

from ..config import settings


DB_DIR = settings.runtime_data_dir / "DB"
TEN_YEARS = DB_DIR / "10 Years Production and Reserves Data.xlsx"
FY_PERF = DB_DIR / "FY2025-26 Perforamance.xlsx"
WORKOVER = DB_DIR / "Workover & Drilling 5 yrs.xlsx"


@dataclass
class YearRow:
    """One row from the 10-Years sheet, cleaned + typed."""
    fy: str                       # "2024-25"
    crude_oil_mmt: float | None
    natural_gas_mmscm: float | None
    p2_oil_reserves_mmt: float | None
    p2_gas_recoverable_bcm: float | None
    p2_remaining_mmtoe: float | None
    reserve_accretion_mmtoe: float | None
    rrr: float | None


@dataclass
class DrillingRow:
    """One row from the FY25-26 drilling annexures."""
    state: str
    target_meterage: float
    target_wells: int
    actual_meterage: float
    actual_wells: int
    pct_achievement: float | None     # 0..1, None if undefined
    behind_wells: int                 # max(0, target_wells - actual_wells)


def _f(v) -> float | None:
    """Coerce a possibly-string '#DIV/0!' Excel value to a clean float or None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.startswith("#") or s == "--":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _i(v) -> int:
    n = _f(v)
    return int(n) if n is not None else 0


# ------------------------------------------------------------------
# 10 Years Production & Reserves
# ------------------------------------------------------------------

@lru_cache(maxsize=1)
def ten_year_rows() -> list[YearRow]:
    wb = openpyxl.load_workbook(str(TEN_YEARS), data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    # First two rows are header/units; year data starts at index 2.
    out: list[YearRow] = []
    for r in rows[2:]:
        fy = (r[0] or "").strip() if isinstance(r[0], str) else None
        if not fy:
            continue
        crude = _f(r[1])
        # Owner-authorised headline correction: FY2025-26 crude is reported as
        # 3.46 MMT (bundled MIS shows 3.45). Mirror of the override in
        # domain_metrics._ten_year_table so the home metrics strip agrees.
        if fy == "2025-26" and crude is not None:
            crude = 3.46
        out.append(YearRow(
            fy=fy,
            crude_oil_mmt=crude,
            natural_gas_mmscm=_f(r[3]),
            p2_oil_reserves_mmt=_f(r[5]),
            p2_gas_recoverable_bcm=_f(r[7]),
            p2_remaining_mmtoe=_f(r[9]),
            reserve_accretion_mmtoe=_f(r[11]),
            rrr=_f(r[13]),
        ))
    return out


def latest_complete_fy() -> str:
    """Most recent FY in the 10-year sheet that has BOTH reserves and RRR
    populated (i.e. not the in-progress current FY)."""
    for r in reversed(ten_year_rows()):
        if r.rrr is not None and r.p2_oil_reserves_mmt is not None:
            return r.fy
    return ten_year_rows()[-1].fy


# ------------------------------------------------------------------
# FY 2025-26 drilling
# ------------------------------------------------------------------

def _read_drilling_sheet(sheet_name: str) -> list[DrillingRow]:
    """Both 'Annexure-III-Expl. Drl' and 'Annexure-IV-Dev. Drl' share the same
    column layout: State, Area, Unit, Tgt-m, Tgt-wells, Act-m, Act-wells, Ach%.
    """
    wb = openpyxl.load_workbook(str(FY_PERF), data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    out: list[DrillingRow] = []
    for r in rows:
        if not r or not r[1]:
            continue
        # Real data rows start with a Round value AND a state OR start with
        # None + state (continuation). We detect by presence of numeric columns.
        target_m = _f(r[4])
        target_w = _f(r[5])
        actual_m = _f(r[6])
        actual_w = _f(r[7])
        if target_m is None and actual_m is None:
            continue
        state = str(r[1]).strip()
        if state.lower().startswith("total"):
            continue
        target_w_i = int(target_w or 0)
        actual_w_i = int(actual_w or 0)
        out.append(DrillingRow(
            state=state,
            target_meterage=target_m or 0.0,
            target_wells=target_w_i,
            actual_meterage=actual_m or 0.0,
            actual_wells=actual_w_i,
            pct_achievement=(actual_m / target_m) if (target_m and target_m > 0) else None,
            behind_wells=max(0, target_w_i - actual_w_i),
        ))
    return out


@lru_cache(maxsize=1)
def exploratory_drilling() -> list[DrillingRow]:
    return _read_drilling_sheet("Annexure-III-Expl. Drl")


@lru_cache(maxsize=1)
def development_drilling() -> list[DrillingRow]:
    return _read_drilling_sheet("Annexure-IV-Dev. Drl")


# ------------------------------------------------------------------
# Production (FY 25-26)
# ------------------------------------------------------------------

@dataclass
class ProductionRow:
    activity: str         # "Crude Oil Production" | "Total ..."
    state: str | None
    unit: str | None
    target: float | None
    actual: float | None
    pct_achievement: float | None


@lru_cache(maxsize=1)
def production_fy25_26() -> list[ProductionRow]:
    wb = openpyxl.load_workbook(str(FY_PERF), data_only=True, read_only=True)
    ws = wb["Annexure-V-Production "]
    rows = list(ws.iter_rows(values_only=True))
    out: list[ProductionRow] = []
    current_activity = ""
    for r in rows[5:]:  # header is rows 0..4
        if not r:
            continue
        act = (r[0] or "").strip() if isinstance(r[0], str) else None
        st = (r[1] or "").strip() if isinstance(r[1], str) else None
        unit = (r[2] or "").strip() if isinstance(r[2], str) else None
        tgt = _f(r[3])
        act_v = _f(r[4])
        ach = _f(r[5])
        if act:
            current_activity = act
        if st is None and act is None:
            continue
        out.append(ProductionRow(
            activity=act or current_activity,
            state=st,
            unit=unit,
            target=tgt,
            actual=act_v,
            pct_achievement=ach,
        ))
    return out
